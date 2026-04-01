"""
export_excel.py — Exports all analytical data into a formatted Excel workbook.
Run: python export/export_excel.py

Output: exports/Marketing_Ecommerce_Analytics.xlsx
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from pipeline.load import get_connection

OUT_DIR = Path(__file__).parent.parent / "exports"
OUT_DIR.mkdir(exist_ok=True)
EXCEL_PATH = OUT_DIR / "Marketing_Ecommerce_Analytics.xlsx"

# ── Colours ───────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT     = "C00000"
GOLD       = "FFD700"
GREY_BG    = "F2F2F2"
WHITE      = "FFFFFF"

def hdr_fill(hex_color):   return PatternFill("solid", fgColor=hex_color)
def hdr_font(bold=True, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def body_font(bold=False, size=10):
    return Font(bold=bold, size=size, name="Calibri", color="000000")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center")
def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row, col_start, col_end, bg=DARK_BLUE):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill    = hdr_fill(bg)
        cell.font    = hdr_font()
        cell.alignment = center()
        cell.border  = thin_border()

def style_data_rows(ws, row_start, row_end, col_start, col_end):
    for r in range(row_start, row_end + 1):
        bg = GREY_BG if r % 2 == 0 else WHITE
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill      = hdr_fill(bg)
            cell.font      = body_font()
            cell.alignment = left()
            cell.border    = thin_border()

def auto_width(ws, min_w=10, max_w=35):
    for col in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

def write_df(ws, df, start_row=2, start_col=1, header_bg=DARK_BLUE):
    # Header
    for ci, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=ci, value=col.replace("_", " ").title())
    style_header_row(ws, start_row, start_col, start_col + len(df.columns) - 1, header_bg)

    # Data
    for ri, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        for ci, val in enumerate(row_data, start=start_col):
            cell = ws.cell(row=ri, column=ci)
            if isinstance(val, float) and np.isnan(val):
                cell.value = None
            else:
                cell.value = val if not isinstance(val, (pd.Timestamp,)) else str(val)

    style_data_rows(ws, start_row + 1, start_row + len(df), start_col, start_col + len(df.columns) - 1)
    return start_row + len(df) + 1  # return next free row


def add_title(ws, title, row=1):
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color=DARK_BLUE, name="Calibri")
    ws.cell(row=row, column=1).alignment = left()
    ws.row_dimensions[row].height = 24


# ── Load data ─────────────────────────────────────────────────
print("Loading data from DuckDB...")
con = get_connection()

tx        = con.execute("SELECT * FROM transactions").df()
customers = con.execute("SELECT * FROM customers").df()
campaigns = con.execute("SELECT * FROM campaigns").df()
products  = con.execute("SELECT * FROM products").df()
events    = con.execute("SELECT * FROM events").df()
c360      = con.execute("SELECT * FROM mart_customer_360").df()
cp        = con.execute("SELECT * FROM mart_campaign_performance").df()
funnel    = con.execute("SELECT * FROM mart_funnel").df()
pp        = con.execute("SELECT * FROM mart_product_performance").df()
con.close()

# Parse dates/derived cols
tx["timestamp"]   = pd.to_datetime(tx["timestamp"])
tx["month"]       = tx["timestamp"].dt.to_period("M").astype(str)
tx["year"]        = tx["timestamp"].dt.year
events["timestamp"] = pd.to_datetime(events["timestamp"])

# ── Build summaries ───────────────────────────────────────────
monthly_rev = (
    tx.groupby("month")
    .agg(orders=("transaction_id","count"),
         gross_revenue=("gross_revenue","sum"),
         net_revenue=("net_revenue","sum"),
         avg_order_value=("net_revenue","mean"),
         refunds=("refund_flag","sum"))
    .reset_index().round(2)
)

channel_rev = (
    tx[tx["campaign_id"] != 0]
    .merge(campaigns[["campaign_id","channel"]].drop_duplicates(), on="campaign_id", how="left")
    .groupby("channel")
    .agg(orders=("transaction_id","count"),
         net_revenue=("net_revenue","sum"),
         unique_customers=("customer_id","nunique"),
         avg_order_value=("net_revenue","mean"))
    .reset_index().round(2).sort_values("net_revenue", ascending=False)
)

rfm_summary = (
    c360[c360["rfm_segment"].notna()]
    .groupby("rfm_segment")
    .agg(customers=("customer_id","count"),
         avg_ltv=("total_revenue","mean"),
         avg_orders=("total_orders","mean"),
         avg_recency_days=("recency_days","mean"))
    .reset_index().round(2).sort_values("avg_ltv", ascending=False)
)

acq_perf = (
    c360[c360["has_purchased"]]
    .groupby("acquisition_channel")
    .agg(customers=("customer_id","count"),
         avg_ltv=("total_revenue","mean"),
         avg_orders=("total_orders","mean"))
    .reset_index().round(2).sort_values("avg_ltv", ascending=False)
)

country_rev = (
    c360[c360["has_purchased"]]
    .groupby("country")
    .agg(customers=("customer_id","count"),
         total_revenue=("total_revenue","sum"),
         avg_ltv=("total_revenue","mean"))
    .reset_index().round(2).sort_values("total_revenue", ascending=False).head(20)
)

cat_perf = (
    pp.groupby("category")
    .agg(products=("product_id","count"),
         total_revenue=("total_revenue","sum"),
         units_sold=("units_sold","sum"),
         avg_price=("base_price","mean"))
    .reset_index().round(2).sort_values("total_revenue", ascending=False)
)

top_products = (
    pp[pp["total_revenue"] > 0]
    .nlargest(25, "total_revenue")
    [[c for c in ["product_id","category","brand","base_price","is_premium",
                  "units_sold","total_revenue","unique_buyers","discount_rate"] if c in pp.columns]]
    .round(3)
)

camp_perf = (
    cp[cp["net_revenue"].notna()]
    [["campaign_id","channel","objective","target_segment","duration_days",
      "tx_count","net_revenue","conversion_rate","bounce_rate","revenue_per_visitor"]]
    .sort_values("net_revenue", ascending=False)
    .round(4)
)

ab_test = (
    events.groupby("experiment_group")
    .agg(sessions=("session_id","nunique"),
         users=("customer_id","nunique"),
         purchases=("event_type", lambda x: (x=="purchase").sum()),
         bounces=("event_type",   lambda x: (x=="bounce").sum()),
         avg_session_sec=("session_duration_sec","mean"))
    .reset_index().round(2)
)
ab_test["conversion_rate"] = (ab_test["purchases"] / ab_test["sessions"]).round(4)
ab_test["bounce_rate"]     = (ab_test["bounces"]   / ab_test["sessions"]).round(4)

funnel_display = funnel.copy()
funnel_display.columns = ["Stage","Count","Drop Off Rate","Conversion From Top"]

# ── Create workbook ───────────────────────────────────────────
print("Building Excel workbook...")
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ═══════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY (KPIs)
# ═══════════════════════════════════════════════
ws = wb.create_sheet("📊 Executive Summary")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 22

add_title(ws, "Marketing & E-Commerce Analytics — Executive Summary", row=1)
ws.row_dimensions[2].height = 6  # spacer

# KPI table header
kpi_header_row = 3
ws.cell(row=kpi_header_row, column=1, value="KPI")
ws.cell(row=kpi_header_row, column=2, value="Value")
ws.cell(row=kpi_header_row, column=3, value="KPI")
ws.cell(row=kpi_header_row, column=4, value="Value")
style_header_row(ws, kpi_header_row, 1, 4, DARK_BLUE)

kpis = [
    ("Total Net Revenue",        f"${tx['net_revenue'].sum():,.0f}"),
    ("Total Orders",             f"{len(tx):,}"),
    ("Unique Customers",         f"{tx['customer_id'].nunique():,}"),
    ("Average Order Value",      f"${tx['net_revenue'].mean():,.2f}"),
    ("Refund Rate",              f"{tx['refund_flag'].mean()*100:.2f}%"),
    ("Total Customers",          f"{len(customers):,}"),
    ("Total Products",           f"{len(products):,}"),
    ("Total Campaigns",          f"{len(campaigns):,}"),
    ("Total Events (2M)",        f"{len(events):,}"),
    ("Avg Campaign Conversion",  f"{cp['conversion_rate'].mean()*100:.2f}%"),
    ("Champions (RFM)",          f"{(c360['rfm_segment']=='Champions').sum():,}"),
    ("Lost Customers (RFM)",     f"{(c360['rfm_segment']=='Lost').sum():,}"),
]
for i, (k, v) in enumerate(kpis):
    r = kpi_header_row + 1 + (i // 2)
    col_k = 1 + (i % 2) * 2
    col_v = col_k + 1
    ws.cell(row=r, column=col_k, value=k).font = body_font(bold=True)
    ws.cell(row=r, column=col_v, value=v).font = body_font()
    for c in [col_k, col_v]:
        ws.cell(row=r, column=c).fill = hdr_fill(GREY_BG if r % 2 == 0 else WHITE)
        ws.cell(row=r, column=c).border = thin_border()
        ws.cell(row=r, column=c).alignment = left()

# ═══════════════════════════════════════════════
# SHEET 2 — REVENUE TRENDS
# ═══════════════════════════════════════════════
ws = wb.create_sheet("📈 Revenue Trends")
ws.sheet_view.showGridLines = False
add_title(ws, "Monthly Revenue & Order Trends")

write_df(ws, monthly_rev, start_row=2)
auto_width(ws)

# Bar chart — orders
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Monthly Orders"
chart1.style = 10
chart1.y_axis.title = "Orders"
chart1.x_axis.title = "Month"
chart1.shape = 4
data_ref = Reference(ws, min_col=3, max_col=3, min_row=2, max_row=2+len(monthly_rev))
cats_ref  = Reference(ws, min_col=1, min_row=3, max_row=2+len(monthly_rev))
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.series[0].graphicalProperties.solidFill = MID_BLUE
chart1.width = 28; chart1.height = 14
ws.add_chart(chart1, f"H2")

# Line chart — revenue
chart2 = LineChart()
chart2.title = "Monthly Net Revenue"
chart2.style = 10
chart2.y_axis.title = "Revenue ($)"
chart2.x_axis.title = "Month"
rev_ref = Reference(ws, min_col=4, max_col=4, min_row=2, max_row=2+len(monthly_rev))
chart2.add_data(rev_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.series[0].graphicalProperties.solidFill   = ACCENT
chart2.series[0].graphicalProperties.line.solidFill = ACCENT
chart2.width = 28; chart2.height = 14
ws.add_chart(chart2, f"H20")

# ═══════════════════════════════════════════════
# SHEET 3 — CUSTOMER ANALYTICS
# ═══════════════════════════════════════════════
ws = wb.create_sheet("👥 Customer Analytics")
ws.sheet_view.showGridLines = False
add_title(ws, "Customer Analytics — RFM, LTV, Acquisition")

row = 2
ws.cell(row=row, column=1, value="RFM Segment Performance").font = hdr_font(color=DARK_BLUE, size=11)
row += 1
row = write_df(ws, rfm_summary, start_row=row, header_bg=DARK_BLUE)

row += 1
ws.cell(row=row, column=1, value="Acquisition Channel Performance").font = hdr_font(color=DARK_BLUE, size=11)
row += 1
row = write_df(ws, acq_perf, start_row=row, header_bg=MID_BLUE)

row += 1
ws.cell(row=row, column=1, value="Revenue by Country (Top 20)").font = hdr_font(color=DARK_BLUE, size=11)
row += 1
write_df(ws, country_rev, start_row=row, header_bg=MID_BLUE)
auto_width(ws)

# Pie chart — RFM segments
pie = PieChart()
pie.title = "RFM Segment Distribution"
pie.style = 10
labels = Reference(ws, min_col=1, min_row=4, max_row=3+len(rfm_summary))
data   = Reference(ws, min_col=2, min_row=3, max_row=3+len(rfm_summary))
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.width = 18; pie.height = 14
ws.add_chart(pie, "H2")

# ═══════════════════════════════════════════════
# SHEET 4 — CAMPAIGN PERFORMANCE
# ═══════════════════════════════════════════════
ws = wb.create_sheet("📣 Campaign Performance")
ws.sheet_view.showGridLines = False
add_title(ws, "Campaign Performance — ROI, Conversion, A/B Test")

row = 2
ws.cell(row=row, column=1, value="Campaign Detail (sorted by Revenue)").font = hdr_font(color=DARK_BLUE, size=11)
row += 1
row = write_df(ws, camp_perf.reset_index(drop=True), start_row=row, header_bg=DARK_BLUE)

row += 1
ws.cell(row=row, column=1, value="Revenue by Channel").font = hdr_font(color=DARK_BLUE, size=11)
row += 1
row = write_df(ws, channel_rev, start_row=row, header_bg=MID_BLUE)

row += 1
ws.cell(row=row, column=1, value="A/B Test — Experiment Groups").font = hdr_font(color=DARK_BLUE, size=11)
row += 1
write_df(ws, ab_test, start_row=row, header_bg=MID_BLUE)
auto_width(ws)

# Bar chart — revenue by channel
ch_chart = BarChart()
ch_chart.type = "col"
ch_chart.title = "Net Revenue by Channel"
ch_chart.style = 10
ch_ch_start = camp_perf.shape[0] + 6
data_ref = Reference(ws, min_col=3, max_col=3,
                     min_row=ch_ch_start, max_row=ch_ch_start + len(channel_rev))
cats_ref  = Reference(ws, min_col=1, min_row=ch_ch_start+1, max_row=ch_ch_start + len(channel_rev))
ch_chart.add_data(data_ref, titles_from_data=True)
ch_chart.set_categories(cats_ref)
ch_chart.series[0].graphicalProperties.solidFill = MID_BLUE
ch_chart.width = 22; ch_chart.height = 14
ws.add_chart(ch_chart, "L2")

# ═══════════════════════════════════════════════
# SHEET 5 — PRODUCT ANALYTICS
# ═══════════════════════════════════════════════
ws = wb.create_sheet("🛍️ Product Analytics")
ws.sheet_view.showGridLines = False
add_title(ws, "Product Analytics — Revenue, Categories, Top Products")

row = 2
ws.cell(row=row, column=1, value="Revenue by Category").font = hdr_font(color=DARK_BLUE, size=11)
row += 1
cat_start = row
row = write_df(ws, cat_perf, start_row=row, header_bg=DARK_BLUE)

row += 1
ws.cell(row=row, column=1, value="Top 25 Products by Revenue").font = hdr_font(color=DARK_BLUE, size=11)
row += 1
write_df(ws, top_products.reset_index(drop=True), start_row=row, header_bg=MID_BLUE)
auto_width(ws)

# Bar chart — category revenue
cat_chart = BarChart()
cat_chart.type = "col"
cat_chart.title = "Revenue by Category"
cat_chart.style = 10
data_ref = Reference(ws, min_col=3, max_col=3, min_row=cat_start, max_row=cat_start+len(cat_perf))
cats_ref  = Reference(ws, min_col=1, min_row=cat_start+1, max_row=cat_start+len(cat_perf))
cat_chart.add_data(data_ref, titles_from_data=True)
cat_chart.set_categories(cats_ref)
cat_chart.series[0].graphicalProperties.solidFill = GOLD
cat_chart.width = 20; cat_chart.height = 14
ws.add_chart(cat_chart, "J2")

# ═══════════════════════════════════════════════
# SHEET 6 — FUNNEL ANALYSIS
# ═══════════════════════════════════════════════
ws = wb.create_sheet("🔽 Funnel Analysis")
ws.sheet_view.showGridLines = False
add_title(ws, "Conversion Funnel Analysis")

row = 2
write_df(ws, funnel_display, start_row=row, header_bg=DARK_BLUE)
auto_width(ws)

# Funnel bar chart
funnel_chart = BarChart()
funnel_chart.type = "bar"  # horizontal
funnel_chart.title = "Conversion Funnel"
funnel_chart.style = 10
data_ref = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=2+4)
cats_ref  = Reference(ws, min_col=1, min_row=3, max_row=2+4)
funnel_chart.add_data(data_ref, titles_from_data=True)
funnel_chart.set_categories(cats_ref)
funnel_chart.series[0].graphicalProperties.solidFill = MID_BLUE
funnel_chart.width = 24; funnel_chart.height = 14
ws.add_chart(funnel_chart, "G2")

# ═══════════════════════════════════════════════
# SHEET 7 — RAW MARTS (for Power BI)
# ═══════════════════════════════════════════════
raw_sheets = {
    "Data_Transactions":  tx[["transaction_id","timestamp","customer_id","product_id",
                               "quantity","discount_applied","gross_revenue","net_revenue",
                               "campaign_id","refund_flag","month","year"]].head(50000),
    "Data_Customers":     customers,
    "Data_Campaigns":     campaigns,
    "Data_Products":      products,
    "Data_Customer360":   c360[["customer_id","country","age","gender","loyalty_tier",
                                 "acquisition_channel","total_orders","total_revenue",
                                 "avg_order_value","recency_days","rfm_score","rfm_segment",
                                 "has_purchased"]],
}
for sheet_name, df in raw_sheets.items():
    ws_raw = wb.create_sheet(sheet_name)
    ws_raw.sheet_view.showGridLines = False
    write_df(ws_raw, df.reset_index(drop=True), start_row=1, header_bg=MID_BLUE)
    auto_width(ws_raw)
    print(f"  Sheet '{sheet_name}': {len(df):,} rows")

# ── Save ─────────────────────────────────────────────────────
wb.save(EXCEL_PATH)
print(f"\n✅ Excel saved: {EXCEL_PATH}")
print(f"   Sheets: {[s.title for s in wb.worksheets]}")
